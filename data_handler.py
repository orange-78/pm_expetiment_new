"""
处理实验结果统计表类 data_handler.py
"""

import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from typing import List, Tuple, Optional, Dict, Callable


class DataManager:
    """
    数据管理类，用于管理实验目录下的模型文件和Excel评估文档
    整合了Excel操作功能
    """
    
    def __init__(self, root_path: str, excel_filename: str = 'evaluation.xlsx', sheet_name: str = 'Sheet1'):
        """
        初始化数据管理器
        
        Args:
            root_path: 根目录路径，如 'root'
            excel_filename: Excel文件名，默认为 'evaluation.xlsx'
            sheet_name: 工作表名称，默认使用活动工作表
        """
        self.root_path = Path(root_path)
        self.excel_path = self.root_path / excel_filename
        self.sheet_name = sheet_name
        
        # 确保根目录存在
        if not self.root_path.exists():
            raise ValueError(f"根目录不存在: {self.root_path}")
        
        # 扫描目录并解析配置（严格排序）
        self.folder_configs = self._scan_directories()
        
        # 如果Excel文件不存在，则生成
        if not self.excel_path.exists():
            self._generate_excel()
        
        # 加载Excel文件
        self.wb = load_workbook(self.excel_path)
        self.sheet = self.wb[sheet_name]
        
        # 建立表头映射
        self._build_header_map()
    
    def _scan_directories(self) -> List[Tuple[int, int, str]]:
        """
        扫描根目录下的所有子文件夹，解析lookback和steps
        严格按照lookback从小到大，lookback相同时按steps从小到大排序
        
        Returns:
            配置列表，每个元素为 (lookback, steps, folder_name)
        """
        configs = []
        
        # 遍历根目录下的所有项
        for item in self.root_path.iterdir():
            if item.is_dir():
                folder_name = item.name
                # 尝试解析文件夹名称
                parts = folder_name.split('_')
                if len(parts) == 2:
                    try:
                        lookback = int(parts[0])
                        steps = int(parts[1])
                        configs.append((lookback, steps, folder_name))
                    except ValueError:
                        # 如果无法转换为整数，跳过该文件夹
                        continue
        
        # 严格排序：先按lookback，再按steps
        configs.sort(key=lambda x: (x[0], x[1]))
        return configs
    
    def _generate_excel(self):
        """
        生成evaluation.xlsx文件，包含lookback和steps列
        """
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        
        # 设置表头
        ws.cell(1, 1, 'lookback')
        ws.cell(1, 2, 'steps')
        
        # 填充数据（已经是排序好的）
        for row_idx, (lookback, steps, _) in enumerate(self.folder_configs, start=2):
            ws.cell(row_idx, 1, lookback)
            ws.cell(row_idx, 2, steps)
        
        # 保存文件
        wb.save(self.excel_path)
        print(f"已生成Excel文件: {self.excel_path}")
    
    def _build_header_map(self):
        """建立表头映射"""
        self.header_map = {}
        for col in range(1, self.sheet.max_column + 1):
            header = self.sheet.cell(1, col).value
            if header:
                self.header_map[header] = col
    
    # ==================== Excel操作方法 ====================
    
    def get_all_headers(self):
        """获取所有列索引名称（表头）"""
        return list(self.header_map.keys())

    def get_row_count(self):
        """获取数据行数（不含表头）"""
        return self.sheet.max_row - 1

    def get_col_count(self):
        """获取列数"""
        return self.sheet.max_column

    def get_column_data(self, col, by_header=True):
        """
        获取指定列的数据（不含表头）
        col: 列名 或 列号
        """
        col_idx = self.header_map[col] if by_header else col
        return [self.sheet.cell(row, col_idx).value for row in range(2, self.sheet.max_row + 1)]

    def get_row_data(self, row_idx):
        """
        获取指定行的数据
        row_idx: 从 0 开始，0 表示 Excel 第2行
        """
        excel_row = row_idx + 2  # Excel 实际行号
        return [self.sheet.cell(excel_row, col).value for col in range(1, self.sheet.max_column + 1)]

    def set_value(self, col, row_idx, new_value, by_header=True):
        """
        修改指定数据单元格的内容
        row_idx: 从 0 开始，0 表示 Excel 第2行
        col: 列索引（表头文字 或 列号）
        """
        excel_row = row_idx + 2
        col_idx = self.header_map[col] if by_header else col
        self.sheet.cell(excel_row, col_idx, new_value)

    def add_empty_column(self, new_header, before_header=None):
        """
        新增一个空列
        new_header: 新列的表头
        before_header: 在指定表头前插入；若为 None，则追加到最后
        """
        if before_header and before_header in self.header_map:
            insert_col_idx = self.header_map[before_header]
            self.sheet.insert_cols(insert_col_idx)
            self.sheet.cell(1, insert_col_idx, new_header)
        else:
            insert_col_idx = self.sheet.max_column + 1
            self.sheet.cell(1, insert_col_idx, new_header)

        self._build_header_map()

    def locate_row_by_keys(self, key_col1, key_col2, key_val1, key_val2):
        """
        用两列的值锁定行, 返回逻辑行号 (从0开始)
        """
        col1_idx = self.header_map[key_col1]
        col2_idx = self.header_map[key_col2]

        rows = []
        for row in range(2, self.sheet.max_row + 1):  # Excel 行号
            c1 = self.sheet.cell(row, col1_idx).value
            c2 = self.sheet.cell(row, col2_idx).value
            if c1 == key_val1 and c2 == key_val2:
                rows.append(row - 2)  # 转换为逻辑行号
        
        return rows

    def modify_cell_by_keys(self, key_col1, key_col2, key_val1, key_val2, target_col, new_value, limit_one=False):
        """
        根据表头文字定位列，再用两列的值锁定行，修改目标列
        """
        target_col_idx = self.header_map[target_col]

        rows = self.locate_row_by_keys(key_col1, key_col2, key_val1, key_val2)
        if limit_one and len(rows) != 1:
            return False
        
        for row_idx in rows:
            excel_row = row_idx + 2
            self.sheet.cell(excel_row, target_col_idx, new_value)
        
        return True
    
    def add_columns_with_data(self, columns_info: List[Dict], data_generator: Callable[[int], Dict[str, any]]):
        """
        批量添加多个列，并按行填充数据
        
        Args:
            columns_info: 列信息列表，每个元素为字典 {'header': '列名', 'before': '插入位置'}
                         before为可选，指定在哪个列之前插入，None表示追加到最后
            data_generator: 数据生成函数，接受行索引(从0开始)，返回字典 {列名: 值}
        
        Example:
            dm.add_columns_with_data(
                [
                    {'header': 'accuracy', 'before': None},
                    {'header': 'loss', 'before': None}
                ],
                lambda row_idx: {
                    'accuracy': calculate_accuracy(row_idx),
                    'loss': calculate_loss(row_idx)
                }
            )
        """
        # 第一步：添加所有空列
        for col_info in columns_info:
            header = col_info['header']
            before = col_info.get('before', None)
            self.add_empty_column(header, before)
        
        # 第二步：逐行填充数据
        row_count = self.get_row_count()
        for row_idx in range(row_count):
            # 调用数据生成函数获取当前行的数据
            row_data = data_generator(row_idx)
            
            # 填充每一列
            for col_info in columns_info:
                header = col_info['header']
                if header in row_data:
                    self.set_value(header, row_idx, row_data[header], by_header=True)
    
    def save(self, new_path=None):
        """保存文件，支持原地或另存为"""
        path = new_path if new_path else self.excel_path
        self.wb.save(path)
    
    # ==================== 模型路径查询方法 ====================
    
    def get_model_paths(self, model_name: str) -> List[str]:
        """
        根据模型名称获取所有配置目录下的模型路径列表
        
        Args:
            model_name: 模型名称，如 'model1'
        
        Returns:
            模型路径列表，如 ['600_100/model1.keras', '600_200/model1.keras']
        """
        model_paths = []
        
        # 如果model_name没有.keras后缀，自动添加
        if not model_name.endswith('.keras'):
            model_filename = f"{model_name}.keras"
        else:
            model_filename = model_name
        
        # 遍历所有配置目录
        for _, _, folder_name in self.folder_configs:
            model_path = f"{self.root_path}/{folder_name}/{model_filename}"
            model_paths.append(model_path)
        
        return model_paths
    
    def get_existing_model_paths(self, model_name: str) -> List[str]:
        """
        根据模型名称获取所有实际存在的模型路径列表
        
        Args:
            model_name: 模型名称，如 'model1'
        
        Returns:
            实际存在的模型路径列表
        """
        model_paths = []
        
        # 如果model_name没有.keras后缀，自动添加
        if not model_name.endswith('.keras'):
            model_filename = f"{model_name}.keras"
        else:
            model_filename = model_name
        
        # 遍历所有配置目录
        for _, _, folder_name in self.folder_configs:
            model_path = f"{self.root_path}/{folder_name}/{model_filename}"
            
            # 只添加实际存在的文件
            if Path(model_path).exists():
                model_paths.append(model_path)
        
        return model_paths
    
    def get_model_paths_with_configs(self, model_name: str) -> List[Tuple[str, int, int]]:
        """
        根据模型名称获取所有配置目录下的模型路径及其配置信息
        
        Args:
            model_name: 模型名称，如 'model1'
        
        Returns:
            元组列表，每个元组为 (模型路径, lookback, steps)
            如 [('600_100/model1.keras', 600, 100), ('600_200/model1.keras', 600, 200)]
        """
        model_info = []
        
        # 如果model_name没有.keras后缀，自动添加
        if not model_name.endswith('.keras'):
            model_filename = f"{model_name}.keras"
        else:
            model_filename = model_name
        
        # 遍历所有配置目录
        for lookback, steps, folder_name in self.folder_configs:
            model_path = f"{self.root_path}/{folder_name}/{model_filename}"
            model_info.append((model_path, lookback, steps))
        
        return model_info
    
    def get_existing_model_paths_with_configs(self, model_name: str) -> List[Tuple[str, int, int]]:
        """
        根据模型名称获取所有实际存在的模型路径及其配置信息
        
        Args:
            model_name: 模型名称，如 'model1'
        
        Returns:
            元组列表，每个元组为 (模型路径, lookback, steps)，仅包含实际存在的文件
        """
        model_info = []
        
        # 如果model_name没有.keras后缀，自动添加
        if not model_name.endswith('.keras'):
            model_filename = f"{model_name}.keras"
        else:
            model_filename = model_name
        
        # 遍历所有配置目录
        for lookback, steps, folder_name in self.folder_configs:
            model_path = f"{self.root_path}/{folder_name}/{model_filename}"
            
            # 只添加实际存在的文件
            if Path(model_path).exists():
                model_info.append((model_path, lookback, steps))
        
        return model_info
    
    def list_all_models_in_folder(self, folder_name: str) -> List[str]:
        """
        列出指定文件夹中的所有.keras模型文件
        
        Args:
            folder_name: 文件夹名称，如 '600_100'
        
        Returns:
            模型文件名列表
        """
        folder_path = self.root_path / folder_name
        
        if not folder_path.exists() or not folder_path.is_dir():
            return []
        
        model_files = []
        for file in folder_path.iterdir():
            if file.is_file() and file.suffix == '.keras':
                model_files.append(file.name)
        
        return sorted(model_files)
    
    def get_all_unique_models(self) -> List[str]:
        """
        获取所有配置目录中的唯一模型名称（不含.keras后缀）
        
        Returns:
            唯一模型名称列表
        """
        all_models = set()
        
        for _, _, folder_name in self.folder_configs:
            models = self.list_all_models_in_folder(folder_name)
            for model in models:
                # 去掉.keras后缀
                model_name = model.replace('.keras', '')
                all_models.add(model_name)
        
        return sorted(list(all_models))
    
    def get_configs(self) -> List[Tuple[int, int, str]]:
        """
        获取所有配置信息
        
        Returns:
            配置列表，每个元素为 (lookback, steps, folder_name)
        """
        return self.folder_configs.copy()
    
    def get_excel_path(self) -> Path:
        """
        获取Excel文件的完整路径
        
        Returns:
            Excel文件路径
        """
        return self.excel_path


# 使用示例
if __name__ == "__main__":
    # 初始化数据管理器
    dm = DataManager('root')
    
    # 查看所有配置（验证排序）
    print("配置列表（已排序）:")
    for lookback, steps, folder in dm.get_configs():
        print(f"  lookback={lookback}, steps={steps}, folder={folder}")
    
    # 获取特定模型的所有路径
    model_paths = dm.get_model_paths('model1')
    print("\nModel1的所有路径:")
    for path in model_paths:
        print(f"  {path}")
    
    # 示例：批量添加列并填充数据
    def generate_row_data(row_idx):
        """为每一行生成数据的函数"""
        row_data = dm.get_row_data(row_idx)
        lookback = row_data[0]  # 假设lookback在第一列
        steps = row_data[1]     # 假设steps在第二列
        
        return {
            'accuracy': 0.95 + row_idx * 0.01,  # 示例数据
            'loss': 0.05 - row_idx * 0.001,     # 示例数据
            'score': lookback * steps / 1000    # 示例计算
        }
    
    # 添加三列并填充数据
    dm.add_columns_with_data(
        [
            {'header': 'accuracy'},
            {'header': 'loss'},
            {'header': 'score'}
        ],
        generate_row_data
    )
    
    # 保存文件
    dm.save()
    print("\n已保存Excel文件")